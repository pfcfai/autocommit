#!/usr/bin/env python
# coding: utf-8



import pandas as pd
'''
df_nt=pd.read_csv('https://raw.githubusercontent.com/laurenttang/pfcfProject/main/Overseas3D/NTD.csv') 
df_nt=pd.read_csv('https://raw.githubusercontent.com/laurenttang/pfcfProject/main/FTDOpenData015.csv')
NTD=df_nt['NTD/USD'].iloc[-1]
print(NTD)
'''


# making list of list 
import openpyxl
from openpyxl.reader.excel import load_workbook
from datetime import timezone
from datetime import datetime

wb=openpyxl.load_workbook('/home/pfcf/pfcfProject/EXCEL_AutoUpdate/FTDOpenData015.xlsx')

print(wb.sheetnames[0])
sheet=wb['FTDOpenData015']
rows=sheet.rows

def recordlist(row):
    #date=sheet['A'+str(i)].value (original code)
    
    datestr=str(row[0].value)
    try:
        dt = datetime.strptime(datestr, "%Y%m%d")
        #dt = datetime(2007, 1, 1)
        datestr = dt.replace(tzinfo=timezone.utc).timestamp()*1000
    except:
        pass
    twse=row[1].value
    

    record=[datestr,twse]
    return record

#mylist=[recordlist(i) for i in range(3,3292)]
mylist = []
innerlist=[]
'''
for row in rows:
    for cell in row:
        innerlist.append(cell.value)
    mylist.append(innerlist)
'''
mylist=[recordlist(row) for row in rows]
del mylist[0]
print(mylist)

# save as json
import json

my_details = {
    'name': 'John Doe',
    'age': 29
}

with open('test.json', 'w') as json_file:
    json.dump(mylist, json_file)
