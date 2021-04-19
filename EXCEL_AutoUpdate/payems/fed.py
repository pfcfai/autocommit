#!/usr/bin/env python
# coding: utf-8

# In[34]:


# making list of list 
import openpyxl
import statistics as stats
from datetime import timezone
from datetime import datetime

book = openpyxl.load_workbook('/home/spark/autocommit/EXCEL_AutoUpdate/payems/經濟數據-FRED增益集.xlsx')

#sheet = book.active
def pairlist(alpha,beta,i):
    data1=sheet[alpha+str(i)].value
    data1=data1.replace(tzinfo=timezone.utc).timestamp()*1000
    data2=sheet[beta+str(i)].value
    record=[data1,data2]    # if you wanna list of list
    return record

# 央行利率
sheet=book['利率-月']
print(sheet.max_row)
rate=[pairlist('A','B',i) for i in range(8,sheet.max_row+1) if pairlist('A','B',i)[1] != '=NA()']
print(rate)
dict=rate
dict2=rate[-120:]

import json

with open('/home/spark/autocommit/EXCEL_AutoUpdate/payems/federal_effective_rate.json', 'w') as json_file:
    json.dump(dict, json_file)
with open('/home/spark/autocommit/EXCEL_AutoUpdate/payems/federal_rate_10yrs.json', 'w') as json_file:
    json.dump(dict2, json_file)

