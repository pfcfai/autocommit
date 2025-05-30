#!/usr/bin/env python
# coding: utf-8

# In[1]:
'''
核心技術:
透過 openpyxl 套件，讀取選定之活頁簿與工作表，
可以寫成 mylist , (which is list of list) 形式後，
以 pandas dataframe 插回資料庫存取
'''

from decouple import config
'''
use .gitignore ,then git will pass the files in .gitignore
use .env to set userID,password as default, to pass through config and save in Variables
'''
#for mysql db use
host = config('host',default='')
user = config('user',default='')
password = config('password',default='')


# In[1]:


import openpyxl
from openpyxl.reader.excel import load_workbook


# In[2]:


wb=openpyxl.load_workbook('/home/targets/autocommit/EXCEL_AutoUpdate/2024改版_綜合日報_VBA 爬蟲.xlsm')


# In[3]:


print(wb.sheetnames[1])
sheet=wb['總紀錄']
# Define the column you want to check (e.g., column A)
column_to_check = 'A'

# Initialize n to 1, the first row
n = 1

# Iterate from the bottom of the sheet to the top to find the last non-empty cell in the specified column
for row in reversed(list(sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1))):

    if row[0].value is not None:
        n = row[0].row
        break

# Update n to the next row
n += 1

# Now, n contains the row number of the next available row in column A
print("n:", n)
# In[4]:


def recordlist(i):
    date=sheet['A'+str(i)].value
    twse=sheet['B'+str(i)].value
    idx =sheet['V'+str(i)].value*100
    PCratio =sheet['AB'+str(i)].value
    record=[date,twse,idx,PCratio]
    return record

mylist=[recordlist(i) for i in range(n-90,n)]
#print(mylist)

print(type(mylist[-1][0]))
# In[5]:


from pandas import DataFrame

df = DataFrame (mylist,columns=['Date','Twse','Index_p','PCratio_p'])
df['id']=df.index+1
print (df)


# In[6]:


import pymysql
from sqlalchemy import create_engine
dbconn2=pymysql.connect(
            host=host,
            database= "cotdatabase",
            user=user,
            password=password,
            port=3306,
            charset='utf8',
            ) 

from sqlalchemy.types import NVARCHAR, Float, Integer
dtypedict = {
'Date': NVARCHAR(length=255),
'Twse': Float(),
'Index_p': Float(),
'PCratio_p': Float(),    
'id': Integer()
}
print(df.dtypes)
engine2 = create_engine('mysql+pymysql://{}:{}@{}:3306/cotdatabase?charset=utf8'.format(user,password,host))
df.to_sql('retailoiindex', engine2, if_exists = 'replace',index=False, dtype=dtypedict)


# In[ ]:




