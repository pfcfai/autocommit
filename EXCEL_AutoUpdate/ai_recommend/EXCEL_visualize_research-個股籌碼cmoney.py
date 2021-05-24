#!/usr/bin/env python
# coding: utf-8

# In[1]:





# making list of list 
import openpyxl
import statistics as stats
from datetime import timezone
from datetime import datetime

book = openpyxl.load_workbook('CMoneyDatabase.xlsm')

#sheet = book.active
def pairlist(a,b,c,d,e,f,g,i):
    data1=sheet[a+str(i)].value
    data2=sheet[b+str(i)].value
    data3=sheet[c+str(i)].value
    data4=sheet[d+str(i)].value
    data5=sheet[e+str(i)].value
    data6=sheet[f+str(i)].value
    data7=sheet[g+str(i)].value
    record=[data1,data2,data3,data4,data5,data6,data7]    # if you wanna list of list
    return record

# 
sheet=book['持股比率']
print(sheet.max_row)
rate=[pairlist('K','L','M','N','O','P','Q',i) for i in range(6,sheet.max_row+1) if pairlist('K','L','M','N','O','P','Q',i)[1] != '=NA()']
print(rate)

import pandas as pd

columns=['symbol','stockname','foreign','trust','dealers','directors','others']
df = pd.DataFrame(rate, columns=columns)


# In[3]:


# -*- coding: utf-8 -*-
import pymysql
import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.engine.url import URL

# dbconnect
dbconn=pymysql.connect(
            host="103.17.9.213",
            database= "webaiuse",
            user="webmysql@actwebdb2",
            password="AIteam168",
            port=3306,
            charset='utf8',
            )  
targetlist=[]
for i in range(3):
    tb=['stronglong1','stronglong2','stronglong3']
    sql='select stocksymbol from {} limit 1;'.format(tb[i])
    generator_df = pd.read_sql(sql=sql,     # mysql query
                                   con=dbconn)  # size you want to fetch each time (we choose 2-years data) 
    targetlist.append(generator_df.iloc[0,0])
print(targetlist)


# In[37]:


''' json form 
  dict1=
  {
    name: 'Share',
    data: [
      { name: '外資', y: df2.iloc[0,2] },
      { name: '投信', y: df2.iloc[0,3] },
      { name: '自營', y: df2.iloc[0,4] },
      { name: '董監', y: df2.iloc[0,5] },
      { name: '其他', y: df2.iloc[0,6] }
    ]
  }
  '''

def jsondict(target): # target as str , ig. '2357'
    #print(df.loc[df['symbol'] == '1101'])
    #df2=df.loc[df['symbol'].isin(targetlist)]
    df2=df.loc[df['symbol'] == target]
    print(df2)
    print(df2.iloc[0,2],df2.iloc[0,3],df2.iloc[0,4],df2.iloc[0,5],df2.iloc[0,6])

    namelist=['外資','投信','自營','董監','其他']

    dict1={}
    datalist=[]

    for i in range(0,5):
        dict={}
        dict['name']=namelist[i]
        dict['y']=df2.iloc[0,i+2]
        datalist.append(dict)
        dict={}
        #print(datalist)
    dict1['name']='Share'
    dict1['data']=datalist

    return dict1

import json
sr=0
finaldict={}
for target in targetlist:
    json_dict = jsondict(target)
    sr+=1
    finaldict['stronglong{}'.format(sr)]=json_dict
    print(json_dict)
    
with open('aistrong_shares_dict.json', 'w') as json_file:
    json.dump(finaldict, json_file)


