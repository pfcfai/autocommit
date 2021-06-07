#!/usr/bin/env python
# coding: utf-8

# In[2]:


# making list of list 
import openpyxl
import statistics as stats
from datetime import timezone
from datetime import datetime
import json

def recordlist1(alpha,i):
    data=sheet1[alpha+str(i)].value
    record=data    # if you wanna only a list
    #record=[data]    # if you wanna list of list
    return record

def recordlist2(alpha,i):
    data=sheet2[alpha+str(i)].value
    record=data    # if you wanna only a list
    #record=[data]    # if you wanna list of list
    return record

book = openpyxl.load_workbook('/home/spark/autocommit/EXCEL_AutoUpdate/CFTC_匯率期貨.xlsm')
comlist=['AUD','CAD','BP','EU','JPY','DX','SWF']

for curency in range(7):
    
    #sheet = book.active
    sheet1=book['{}1'.format(str(comlist[curency]))]
    sheet2=book['{}2'.format(str(comlist[curency]))]

    print('=== printing {} ==='.format(str(comlist[curency])),sheet1.max_row)

    date=[recordlist1('C',i) for i in range(2,54) if recordlist1('C',i) != None]
    date=[each.strftime("%m/%d") for each in date]
    date.reverse()
    print(date[:])

    price=[round(recordlist1('EA',i),2) for i in range(2,54) if recordlist1('EA',i) != None]
    price.reverse()
    print(price[:])

    spec_long=[recordlist1('I',i) for i in range(2,54) if recordlist1('I',i) != None]
    print(spec_long[:])
    spec_short=[recordlist1('J',i) for i in range(2,54) if recordlist1('J',i) != None]
    print(spec_short[:])

    Net_Spec=[spec_long[i]-spec_short[i] for i in range(52)]
    Net_Spec.reverse()
    print(Net_Spec)

    commercial_long=[recordlist2('O',i) for i in range(2,54) if recordlist2('O',i) != None]
    print(commercial_long[:])
    commercial_short=[recordlist2('P',i) for i in range(2,54) if recordlist2('P',i) != None]
    print(commercial_short[:])

    Net_Commercial=[commercial_long[i]-commercial_short[i] for i in range(52)]
    Net_Commercial.reverse()
    print(Net_Commercial)


    #print(sheet.max_row)
    #print(len(sheet['G']))


    dict={}
    dict['date']=date[-52:]
    dict['price']=price[-52:]
    dict['commercial']=Net_Commercial[-52:]
    dict['spect']=Net_Spec[-52:]
    print(dict)

    with open('/home/spark/autocommit/EXCEL_AutoUpdate/CFTC_{}.json'.format(str(comlist[curency])), 'w') as json_file:
        json.dump(dict, json_file)


# In[2]:


# making list of list 
import openpyxl
import statistics as stats
from datetime import timezone
from datetime import datetime
import json

def recordlist1(alpha,i):
    data=sheet1[alpha+str(i)].value
    record=data    # if you wanna only a list
    #record=[data]    # if you wanna list of list
    return record

def recordlist2(alpha,i):
    data=sheet2[alpha+str(i)].value
    record=data    # if you wanna only a list
    #record=[data]    # if you wanna list of list
    return record

book = openpyxl.load_workbook('/home/spark/autocommit/EXCEL_AutoUpdate/CFTC_股債期貨.xlsm')
comlist=['SP','NQ','DJ','US','TY']
priceindex=['DX','DY','DY','EB','EB']

for curency in range(5):
    
    #sheet = book.active
    sheet1=book['{}1'.format(str(comlist[curency]))]
    sheet2=book['{}2'.format(str(comlist[curency]))]

    print('=== printing {} ==='.format(str(comlist[curency])),sheet1.max_row)

    date=[recordlist1('C',i) for i in range(2,54) if recordlist1('C',i) != None]
    date=[each.strftime("%m/%d") for each in date]
    date.reverse()
    print(date[:])

    price=[round(recordlist1(str(priceindex[curency]),i),2) for i in range(2,54) if recordlist1(str(priceindex[curency]),i) != None]
    price.reverse()
    print(price[:])

    spec_long=[recordlist1('I',i) for i in range(2,54) if recordlist1('I',i) != None]
    print(spec_long[:])
    spec_short=[recordlist1('J',i) for i in range(2,54) if recordlist1('J',i) != None]
    print(spec_short[:])

    Net_Spec=[spec_long[i]-spec_short[i] for i in range(52)]
    Net_Spec.reverse()
    print(Net_Spec)

    commercial_long=[recordlist2('O',i) for i in range(2,54) if recordlist2('O',i) != None]
    print(commercial_long[:])
    commercial_short=[recordlist2('P',i) for i in range(2,54) if recordlist2('P',i) != None]
    print(commercial_short[:])

    Net_Commercial=[commercial_long[i]-commercial_short[i] for i in range(52)]
    Net_Commercial.reverse()
    print(Net_Commercial)


    #print(sheet.max_row)
    #print(len(sheet['G']))


    dict={}
    dict['date']=date[-52:]
    dict['price']=price[-52:]
    dict['commercial']=Net_Commercial[-52:]
    dict['spect']=Net_Spec[-52:]
    print(dict)

    with open('/home/spark/autocommit/EXCEL_AutoUpdate/CFTC_{}.json'.format(str(comlist[curency])), 'w') as json_file:
        json.dump(dict, json_file)


# In[4]:


# making list of list 
import openpyxl
import statistics as stats
from datetime import timezone
from datetime import datetime
import json

def recordlist1(alpha,i):
    data=sheet1[alpha+str(i)].value
    record=data    # if you wanna only a list
    #record=[data]    # if you wanna list of list
    return record

def recordlist2(alpha,i):
    data=sheet2[alpha+str(i)].value
    record=data    # if you wanna only a list
    #record=[data]    # if you wanna list of list
    return record

book = openpyxl.load_workbook('/home/spark/autocommit/EXCEL_AutoUpdate/CFTC_農油金期貨.xlsm')
comlist=['soy','wheat','corn','crude','gold']
priceindex=['EB','EB','EB','EA','EA']
for curency in range(5):
    
    #sheet = book.active
    sheet1=book['{}1'.format(str(comlist[curency]))]
    sheet2=book['{}2'.format(str(comlist[curency]))]

    print('=== printing {} ==='.format(str(comlist[curency])),sheet1.max_row)

    date=[recordlist1('C',i) for i in range(2,54) if recordlist1('C',i) != None]
    date=[each.strftime("%m/%d") for each in date]
    date.reverse()
    print(date[:])

    price=[round(recordlist1(str(priceindex[curency]),i),2) for i in range(2,54) if recordlist1(str(priceindex[curency]),i) != None]
    price.reverse()
    print(price[:])

    spec_long=[recordlist1('I',i) for i in range(2,54) if recordlist1('I',i) != None]
    print(spec_long[:])
    spec_short=[recordlist1('J',i) for i in range(2,54) if recordlist1('J',i) != None]
    print(spec_short[:])

    Net_Spec=[spec_long[i]-spec_short[i] for i in range(52)]
    Net_Spec.reverse()
    print(Net_Spec)

    commercial_long=[recordlist2('N',i) for i in range(2,54) if recordlist2('N',i) != None]
    print(commercial_long[:])
    commercial_short=[recordlist2('O',i) for i in range(2,54) if recordlist2('O',i) != None]
    print(commercial_short[:])

    Net_Commercial=[commercial_long[i]-commercial_short[i] for i in range(52)]
    Net_Commercial.reverse()
    print(Net_Commercial)


    #print(sheet.max_row)
    #print(len(sheet['G']))


    dict={}
    dict['date']=date[-52:]
    dict['price']=price[-52:]
    dict['commercial']=Net_Commercial[-52:]
    dict['spect']=Net_Spec[-52:]
    print(dict)

    with open('/home/spark/autocommit/EXCEL_AutoUpdate/CFTC_{}.json'.format(str(comlist[curency])), 'w') as json_file:
        json.dump(dict, json_file)


# In[ ]:




