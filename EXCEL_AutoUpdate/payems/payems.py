#!/usr/bin/env python
# coding: utf-8

# In[1]:


# making list of list 
import openpyxl
import statistics as stats
from datetime import timezone
from datetime import datetime

book = openpyxl.load_workbook('經濟數據-FRED增益集.xlsx')

#sheet = book.active
sheet=book['就業-月']

def recordlist(alpha,i):
    data=sheet[alpha+str(i)].value
    record=data    # if you wanna only a list
    #record=[data]    # if you wanna list of list
    return record

date=[recordlist('A',i) for i in range(8,sheet.max_row+1) if recordlist('A',i) != None]
date=[each.strftime("%m/'%y") for each in date]
print(date[-12:])
unrate=[recordlist('E',i) for i in range(8,sheet.max_row+1) if recordlist('E',i) != None]
print(unrate[-12:])
payems=[recordlist('B',i) for i in range(8,sheet.max_row+1) if recordlist('B',i) != None]
payems_icz=[payems[j]-payems[j-1] for j in range(len(payems)) if j>0]
print(payems_icz[-12:])
#print(sheet.max_row)
#print(len(sheet['G']))


# In[47]:


import json

dict={}
dict['date']=date[-12:]
dict['unrate']=unrate[-12:]
dict['payems_icz']=payems_icz[-12:]
print(dict)

with open('payems.json', 'w') as json_file:
    json.dump(dict, json_file)


# In[4]:


# making list of list 
import openpyxl
import statistics as stats
from datetime import timezone
from datetime import datetime

book = openpyxl.load_workbook('經濟數據-FRED增益集.xlsx')

#sheet = book.active
sheet=book['薪資-月']

def recordlist(alpha,i):
    data=sheet[alpha+str(i)].value
    record=data    # if you wanna only a list
    #record=[data]    # if you wanna list of list
    return record

date=[recordlist('A',i) for i in range(8,sheet.max_row+1) if recordlist('A',i) != None]
date=[each.strftime("%m/'%y") for each in date]
print(date[-12:])
hourpay_avg=[recordlist('B',i) for i in range(8,sheet.max_row+1) if recordlist('B',i) != None]
print(hourpay_avg[-13:])

hourpay_yoy=[round((hourpay_avg[j]/hourpay_avg[j-12]-1)*100,1) for j in range(len(hourpay_avg)) if j>12]
print(hourpay_yoy[-12:])
#print(sheet.max_row)
#print(len(sheet['G']))


# In[5]:


import json

dict={}
dict['date']=date[-12:]
dict['hourpay_avg']=hourpay_avg[-12:]
dict['hourpay_yoy']=hourpay_yoy[-12:]
print(dict)

with open('hourpay.json', 'w') as json_file:
    json.dump(dict, json_file)


# In[10]:


# making list of list 
import openpyxl
import statistics as stats
from datetime import timezone
from datetime import datetime

book = openpyxl.load_workbook('經濟數據-FRED增益集.xlsx')

#sheet = book.active
sheet=book['就業-週']

def recordlist(alpha,i):
    data=sheet[alpha+str(i)].value
    record=data    # if you wanna only a list
    #record=[data]    # if you wanna list of list
    return record

date=[recordlist('A',i) for i in range(8,sheet.max_row+1) if recordlist('A',i) != None]
date=[each.strftime("%m/%d") for each in date]
print(date[-12:])
initial_claim=[recordlist('B',i)/1000 for i in range(8,sheet.max_row+1) if recordlist('B',i) != None]
print(initial_claim[-12:])
continued_claim=[recordlist('E',i)/1000 for i in range(8,sheet.max_row+1) if recordlist('E',i) != None]
print(continued_claim[-12:])
#print(sheet.max_row)
#print(len(sheet['G']))


# In[11]:


import json

dict={}
dict['date']=date[-12:]
dict['initial_claim']=initial_claim[-12:]
dict['continued_claim']=continued_claim[-12:]
print(dict)

with open('initial_claim.json', 'w') as json_file:
    json.dump(dict, json_file)


# In[16]:


# making list of list 
import openpyxl
import statistics as stats
from datetime import timezone
from datetime import datetime

book = openpyxl.load_workbook('經濟數據-FRED增益集.xlsx')

#sheet = book.active
def pairlist(alpha,beta,i):
    data1=sheet[alpha+str(i)].value
    data1=data1.replace(tzinfo=timezone.utc).timestamp()*1000
    data2=sheet[beta+str(i)].value/1000
    record=[data1,data2]    # if you wanna list of list
    return record

# PCE
sheet=book['就業-週']
initial_claim   = [pairlist('A','B',i) for i in range(8,sheet.max_row+1) if pairlist('A','B',i)[1] != '=NA()']
continued_claim = [pairlist('D','E',i) for i in range(8,sheet.max_row+1-1) if pairlist('D','E',i)[1] != '=NA()']

dict=initial_claim[-252:]
dict2=continued_claim[-252:]

import json

with open('initial_claim.json', 'w') as json_file:
    json.dump(dict, json_file)
with open('continued_claim.json', 'w') as json_file:
    json.dump(dict2, json_file)



