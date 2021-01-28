#!/usr/bin/env python
# coding: utf-8

# In[1]:


get_ipython().system('pip install openpyxl')


# In[1]:


import openpyxl
from openpyxl.reader.excel import load_workbook


# In[2]:


wb=openpyxl.load_workbook('9-1自動化更新資料庫-108.11.05.xlsm')


# In[3]:


print(wb.sheetnames[1])
sheet=wb['近90天Data']


# In[4]:


def recordlist(i):
    date=sheet['A'+str(i)].value
    twse=sheet['B'+str(i)].value
    idx =sheet['N'+str(i)].value*100
    PCratio =sheet['Q'+str(i)].value
    record=[date,twse,idx,PCratio]
    return record

mylist=[recordlist(i) for i in range(3,93)]
#print(mylist)


# In[5]:


from pandas import DataFrame

df = DataFrame (mylist,columns=['Date','Twse','Index_p','PCratio_p'])
df['id']=df.index+1
print (df)


# In[6]:


import pymysql
from sqlalchemy import create_engine
dbconn2=pymysql.connect(
            host="103.17.9.213",
            database= "cotdatabase",
            user="webmysql@actwebdb2",
            password="AIteam168",
            port=3306,
            charset='utf8',
            ) 

from sqlalchemy.types import NVARCHAR, Float, Integer
dtypedict = {
'Date': NVARCHAR(length=255),
'Twse': Float(),
'Index_p': Float(),
'PCratio_p': Float(),    
'id': Integer()
}
print(df.dtypes)
engine2 = create_engine('mysql+pymysql://webmysql@actwebdb2:AIteam168@103.17.9.213:3306/cotdatabase?charset=utf8')
df.to_sql('retailoiindex', engine2, if_exists = 'replace',index=False, dtype=dtypedict)


# In[ ]:




