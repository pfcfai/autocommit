#!/usr/bin/env python
# coding: utf-8

# In[10]:


# making list of list 
import openpyxl
import statistics as stats
from datetime import timezone
from datetime import datetime

book = openpyxl.load_workbook('經濟數據-FRED增益集.xlsx')

#sheet = book.active
sheet=book['通膨-月']

def recordlist(alpha,i):
    data=sheet[alpha+str(i)].value
    record=data    # if you wanna only a list
    #record=[data]    # if you wanna list of list
    return record

# PCE
date=[recordlist('A',i) for i in range(8,sheet.max_row+1) if recordlist('A',i) != None]
date=[each.strftime("%m/%y") for each in date]
print(date[-12:])

pce=[recordlist('B',i) for i in range(8,sheet.max_row+1) if recordlist('B',i) != None]
pce_yoy=[round((pce[j]/pce[j-12]-1)*100,2) for j in range(len(pce)) if j>12]
print(pce_yoy[-12:])

pcecr=[recordlist('E',i) for i in range(8,sheet.max_row+1) if recordlist('E',i) != None]
pcecr_yoy=[round((pcecr[j]/pcecr[j-12]-1)*100,2) for j in range(len(pcecr)) if j>12]
print(pcecr_yoy[-12:])

# CPI
date2=[recordlist('G',i) for i in range(8,sheet.max_row+1) if recordlist('G',i) != None]
date2=[each.strftime("%m/%y") for each in date2]
print(date2[-12:])

cpi=[recordlist('H',i) for i in range(8,sheet.max_row+1) if recordlist('H',i) != None]
cpi_yoy=[round((cpi[j]/cpi[j-12]-1)*100,2) for j in range(len(cpi)) if j>12]
print(cpi_yoy[-12:])

cpicr=[recordlist('K',i) for i in range(8,sheet.max_row+1) if recordlist('K',i) != None]
cpicr_yoy=[round((cpicr[j]/cpicr[j-12]-1)*100,2) for j in range(len(cpicr)) if j>12]
print(cpicr_yoy[-12:])


# In[11]:


import json

dict={}
dict['date']=date[-12:]
dict['pce_yoy']=pce_yoy[-12:]
dict['pcecr_yoy']=pcecr_yoy[-12:]
print(dict)

dict2={}
dict2['date2']=date2[-12:]
dict2['cpi_yoy']=cpi_yoy[-12:]
dict2['cpicr_yoy']=cpicr_yoy[-12:]
print(dict2)

with open('/home/spark/autocommit/EXCEL_AutoUpdate/payems/pce.json', 'w') as json_file:
    json.dump(dict, json_file)
with open('/home/spark/autocommit/EXCEL_AutoUpdate/payems/cpi.json', 'w') as json_file:
    json.dump(dict2, json_file)


# In[31]:


# making list of list 
import openpyxl
import statistics as stats
from datetime import timezone
from datetime import datetime

book = openpyxl.load_workbook('經濟數據-FRED增益集.xlsx')

#sheet = book.active
def pairlist(alpha,beta,i):
    data1=sheet[alpha+str(i)].value
    data1=data1.replace(tzinfo=timezone.utc).timestamp()*1000
    data2=sheet[beta+str(i)].value
    record=[data1,data2]    # if you wanna list of list
    return record

# PCE
sheet=book['通膨-日']
be_rate=[pairlist('D','E',i) for i in range(8,sheet.max_row+1) if pairlist('D','E',i)[1] != '=NA()']
sheet=book['利率-日']
print(sheet.max_row)
ytm=[pairlist('G','H',i) for i in range(8,sheet.max_row+1) if pairlist('G','H',i)[1] != '=NA()']
print(be_rate[-253:-1])
print(ytm[-252:])
dict=be_rate[-252:]
dict2=ytm[-252:]

import json

with open('/home/spark/autocommit/EXCEL_AutoUpdate/payems/breakeven_rate.json', 'w') as json_file:
    json.dump(dict, json_file)
with open('/home/spark/autocommit/EXCEL_AutoUpdate/payems/10yr_yield.json', 'w') as json_file:
    json.dump(dict2, json_file)

